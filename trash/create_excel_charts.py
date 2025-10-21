import json
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, Alignment, PatternFill

# Load the data
with open('project_site/battle_results.json', 'r') as f:
    battle_results = json.load(f)

with open('project_site/ladder_results.json', 'r') as f:
    ladder_results = json.load(f)

# Create workbook
wb = Workbook()

# ====================
# Sheet 1: Overall Win Rates
# ====================
ws1 = wb.active
ws1.title = "Overall Win Rates"

# Add headers
ws1['A1'] = "Bot Name"
ws1['B1'] = "Win Rate (%)"
ws1['A1'].font = Font(bold=True, size=12)
ws1['B1'].font = Font(bold=True, size=12)

# Calculate overall win rates
random_wins = battle_results['random_vs_maxdamage']['p1_wins'] + battle_results['random_vs_custom']['p1_wins']
random_total = 100  # 50 + 50 battles
random_win_rate = (random_wins / random_total) * 100

maxdamage_wins = battle_results['random_vs_maxdamage']['p2_wins'] + battle_results['maxdamage_vs_custom']['p1_wins']
maxdamage_total = 100
maxdamage_win_rate = (maxdamage_wins / maxdamage_total) * 100

custom_wins = battle_results['random_vs_custom']['p2_wins'] + battle_results['maxdamage_vs_custom']['p2_wins']
custom_total = 100
custom_win_rate = (custom_wins / custom_total) * 100

# Add data
ws1['A2'] = "Random Bot"
ws1['B2'] = random_win_rate
ws1['A3'] = "Max Damage Bot"
ws1['B3'] = maxdamage_win_rate
ws1['A4'] = "Custom Strategy Bot"
ws1['B4'] = custom_win_rate

# Format percentages
for row in range(2, 5):
    ws1[f'B{row}'].number_format = '0.0'

# Adjust column widths
ws1.column_dimensions['A'].width = 25
ws1.column_dimensions['B'].width = 15

# Create chart
chart1 = BarChart()
chart1.type = "bar"
chart1.title = "Overall Win Rate Comparison"
chart1.y_axis.title = "Bot"
chart1.x_axis.title = "Win Rate (%)"
chart1.height = 10
chart1.width = 20

data = Reference(ws1, min_col=2, min_row=1, max_row=4)
cats = Reference(ws1, min_col=1, min_row=2, max_row=4)
chart1.add_data(data, titles_from_data=True)
chart1.set_categories(cats)
chart1.legend = None

ws1.add_chart(chart1, "D2")

# ====================
# Sheet 2: Head-to-Head Results
# ====================
ws2 = wb.create_sheet("Head-to-Head")

# Add headers
ws2['A1'] = "Matchup"
ws2['B1'] = "Bot 1 Wins"
ws2['C1'] = "Bot 2 Wins"
for cell in ['A1', 'B1', 'C1']:
    ws2[cell].font = Font(bold=True, size=12)

# Add data
ws2['A2'] = "Random vs Max Damage"
ws2['B2'] = battle_results['random_vs_maxdamage']['p1_wins']
ws2['C2'] = battle_results['random_vs_maxdamage']['p2_wins']

ws2['A3'] = "Random vs Custom"
ws2['B3'] = battle_results['random_vs_custom']['p1_wins']
ws2['C3'] = battle_results['random_vs_custom']['p2_wins']

ws2['A4'] = "Max Damage vs Custom"
ws2['B4'] = battle_results['maxdamage_vs_custom']['p1_wins']
ws2['C4'] = battle_results['maxdamage_vs_custom']['p2_wins']

# Adjust column widths
ws2.column_dimensions['A'].width = 25
ws2.column_dimensions['B'].width = 15
ws2.column_dimensions['C'].width = 15

# Create stacked bar chart
chart2 = BarChart()
chart2.type = "bar"
chart2.grouping = "stacked"
chart2.title = "Head-to-Head Battle Results (50 battles each)"
chart2.y_axis.title = "Matchup"
chart2.x_axis.title = "Number of Wins"
chart2.height = 10
chart2.width = 20

data = Reference(ws2, min_col=2, min_row=1, max_col=3, max_row=4)
cats = Reference(ws2, min_col=1, min_row=2, max_row=4)
chart2.add_data(data, titles_from_data=True)
chart2.set_categories(cats)

ws2.add_chart(chart2, "E2")

# ====================
# Sheet 3: Ladder Performance
# ====================
ws3 = wb.create_sheet("Ladder Performance")

# Add headers
ws3['A1'] = "Version"
ws3['B1'] = "Wins"
ws3['C1'] = "Losses"
ws3['D1'] = "Win Rate (%)"
for cell in ['A1', 'B1', 'C1', 'D1']:
    ws3[cell].font = Font(bold=True, size=12)

# Add data
ws3['A2'] = "Custom Strategy v1 (Basic)"
ws3['B2'] = ladder_results['custom_strategy_v1']['wins']
ws3['C2'] = ladder_results['custom_strategy_v1']['losses']
ws3['D2'] = ladder_results['custom_strategy_v1']['win_rate']

ws3['A3'] = "Custom Strategy v2 (Improved)"
ws3['B3'] = ladder_results['custom_strategy_v2']['wins']
ws3['C3'] = ladder_results['custom_strategy_v2']['losses']
ws3['D3'] = ladder_results['custom_strategy_v2']['win_rate']

# Format percentages
ws3['D2'].number_format = '0.0'
ws3['D3'].number_format = '0.0'

# Adjust column widths
ws3.column_dimensions['A'].width = 30
ws3.column_dimensions['B'].width = 12
ws3.column_dimensions['C'].width = 12
ws3.column_dimensions['D'].width = 15

# Create stacked bar chart for wins/losses
chart3 = BarChart()
chart3.type = "bar"
chart3.grouping = "stacked"
chart3.title = "Ladder Performance: v1 vs v2"
chart3.y_axis.title = "Version"
chart3.x_axis.title = "Number of Battles"
chart3.height = 10
chart3.width = 20

data = Reference(ws3, min_col=2, min_row=1, max_col=3, max_row=3)
cats = Reference(ws3, min_col=1, min_row=2, max_row=3)
chart3.add_data(data, titles_from_data=True)
chart3.set_categories(cats)

ws3.add_chart(chart3, "F2")

# Create win rate comparison chart
chart4 = BarChart()
chart4.type = "bar"
chart4.title = "Win Rate Comparison"
chart4.y_axis.title = "Version"
chart4.x_axis.title = "Win Rate (%)"
chart4.height = 10
chart4.width = 20

data = Reference(ws3, min_col=4, min_row=1, max_row=3)
cats = Reference(ws3, min_col=1, min_row=2, max_row=3)
chart4.add_data(data, titles_from_data=True)
chart4.set_categories(cats)
chart4.legend = None

ws3.add_chart(chart4, "F18")

# Add summary text
ws3['A6'] = "Summary:"
ws3['A6'].font = Font(bold=True, size=12)
ws3['A7'] = f"Improvement: {ladder_results['custom_strategy_v2']['win_rate'] - ladder_results['custom_strategy_v1']['win_rate']}% increase in win rate"
ws3['A8'] = "The v2 improvements included better Dynamax timing, setup moves, and switching logic."

# Save workbook
wb.save('Pokemon_Bot_Performance_Analysis.xlsx')
print("Excel file created successfully: Pokemon_Bot_Performance_Analysis.xlsx")
